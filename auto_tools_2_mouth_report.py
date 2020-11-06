'''
Author: John Diamond
Date: 2020-10-19 10:12:38
LastEditors: John Diamond
LastEditTime: 2020-11-06 17:16:26
FilePath: \auto-tools\aotu_tools_2_mouth_report.py
'''
# -*- coding:utf-8 -*-
#import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

#from datetime import date
#from datetime import datetime

class cFileOpt:

	def __init__(self):
		self.InputFil = load_workbook('mouth.xlsx') 			# data_only=True 表示读取表格中公式的计算结果 不读取公式
		print('文件 mouth.xlsx 打开成功！')
		print()
		print('输入要操作的sheet的位置数，比如第一个sheet就输入 1 然后回车，其他sheet以此类推...\n')
		num_sheet = input('要操作的sheet数：')
		print('输入 本月合计 的所在列，如AA列 输入AA 然后回车即可...\n')
		self.Tmpcolnum_mouth = input('本月合计列：')
		if len(self.Tmpcolnum_mouth) == 1:
			self.colnum_mouth = ord(self.Tmpcolnum_mouth[0]) - 64
			#print(ord(self.Tmpcolnum_mouth[0]))
		elif len(self.Tmpcolnum_mouth) == 2:
			self.colnum_mouth = (ord(self.Tmpcolnum_mouth[0]) - 39) + ord(self.Tmpcolnum_mouth[1]) - 64
			#print(ord(self.Tmpcolnum_mouth[0]), ord(self.Tmpcolnum_mouth[1]))
		#print(self.colnum_mouth)
		self.Mouth_Sht = self.InputFil.worksheets[int(num_sheet) - 1]					# 0、1、2 -- mouth report sheet		4 -- for debug
		pass

	def CloseFiles(self):
		self.InputFil.save(filename = 'mouth.xlsx')
		self.InputFil.close()
		print('文件 mouth.xlsx 保存成功！' )
		pass

	def Prepare(self):

		'''要填充位置的格式内容预处理（清空）'''
		m_list = self.Mouth_Sht.merged_cells							# 取消单元格合并
		cr = []															# 先获取已合并的所有单元格
		for m_area in m_list:
																		# 合并单元格的起始坐标、终止坐标
			r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
			if c1 == c2 == ((self.colnum_mouth + 2)) and r1 != 1:			# 挑选出自己想操作的合并单元格的范围
				cr.append((r1, r2, c1, c2))
		for r in cr:
				self.Mouth_Sht.unmerge_cells(start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])

		for x in range(3, self.Mouth_Sht.max_row+1):
			for y in range((self.colnum_mouth + 2), (self.colnum_mouth + 5)):
				self.Mouth_Sht.cell(row = x,column = y).value = ''		# 清空程序要填写的地方
				self.Mouth_Sht.cell(row = x,column = y).alignment = Alignment(horizontal="center", vertical="center")
																		# 横纵居中
				self.Mouth_Sht.cell(row = x,column = y).font = Font(name='微软雅黑', size=9)
																		# 设置字体 字号
				self.cellborder(x, 0)									# 第二个参数为0 清除边框
				self.Mouth_Sht.cell(row = x,column = y).fill = PatternFill(fill_type=None)	
																		# 去除颜色填充
				pass
			
		self.Mouth_Sht.freeze_panes = 'A3'								# 冻结前两行
		pass

	def ParseRawDat(self):
		ProjectSet = set()												# 去重用的项目名称集合
		ProjectList = []												# 去重之后的项目名称
		ProjectNum = 1													# 项目序号
		RowFillproject = 3												# 略过表头两行
		
		for i in range(3, self.Mouth_Sht.max_row+1, 10):
			ProjectSet.add(self.Mouth_Sht.cell(row = i,column = 1).value)
		ProjectList = list(ProjectSet)									# 项目集合 ——> list
		ProjectList.remove(None)										# 排除无内容列表
		print(ProjectList)
		
		for j in range(len(ProjectList)):								# 项目循环			
			Cnt_matchNum = 0
			SaveCoordMouth = ['','','','','','','','','','']			# 保存N个坐标
			TmpCoordMouth = [[],[],[],[],[],[],[],[],[],[]]				# 临时保存坐标的总列表
			SaveFormulaMouth = []										# 保存十个求和公式 一个循环写进去
			print('No.%d' % ProjectNum)
			print(ProjectList[j])
			self.FillMouthSht((RowFillproject + j * 10), (self.colnum_mouth +2), ProjectList[j]) # 写项目名
			self.cellborder((RowFillproject + j * 10))					# 合并单元格之前 把所有的格格框线画好
			self.mergecells((RowFillproject + j * 10))					# 项目名的10个单元格合并
			self.FillMouthSht((RowFillproject + j * 10), (self.colnum_mouth + 3), 0)	# 写执行阶段
			
			for projectrow in range(3, self.Mouth_Sht.max_row+1, 10):	# 遍历以查找
				#print(projectrow)
				CelVal = self.Mouth_Sht.cell(row = projectrow,column = 1).value
				if CelVal == ProjectList[j]:							# 符合项目名称
					for phaseloop in range(0,10):
						if self.Mouth_Sht.cell(row = projectrow + phaseloop,column = self.colnum_mouth).value == None:
							continue
							pass
						else:
							TmpCoordMouth[phaseloop].append('%s%d,' % (self.Tmpcolnum_mouth, (projectrow + phaseloop)))
						#CntPhaseList[phaseloop] += self.Mouth_Sht.cell(row = projectrow + phaseloop,column = self.colnum_mouth).value # 读公式结果的实现
					Cnt_matchNum += 1									# 与当前项目匹配的岗位个数
					pass

			for i in range(0, 10):										# 把当前项目的所有岗位的十个阶段的数量填到月总合计里
				for saveloop in range(len(TmpCoordMouth[i])):			# 把当前项目的所有岗位的十个阶段的数量填到月总合计里
					SaveCoordMouth[i] += TmpCoordMouth[i][saveloop]
				pass
			for fillloop in range(0, 10):								# 把当前项目的所有岗位的十个阶段的数量填到月总合计里
				#self.FillMouthSht((RowFillproject + j * 10 + fillloop), 33, CntPhaseList[fillloop]) # 读公式结果的实现
				SaveFormulaMouth.append('=SUM(%s)' % SaveCoordMouth[fillloop])
				self.FillMouthSht((RowFillproject + j * 10 + fillloop), (self.colnum_mouth + 4), SaveFormulaMouth[fillloop])
				pass
			print('当前项目有%d个岗位' % Cnt_matchNum)
			print('\n')
			ProjectNum += 1												# 完成一个项目 序号+1
			pass

	def FillMouthSht(self, i, j, celval):
		phase = ['需求','推荐简历','有效简历数','一面（到面）','终面','offer','入职','转入','在职','离职']
		if j == (self.colnum_mouth + 3):
			for x in range(0, 10):
				self.Mouth_Sht.cell(row = (i + x),column = (self.colnum_mouth + 3)).value = phase[x]
				if x == 0 or x == 3 or x == 6 or x == 9:
					self.Mouth_Sht.cell(row = (i + x),column = (self.colnum_mouth + 3)).fill = PatternFill(start_color ='FFFF00', end_color = 'FFFF00', fill_type = 'solid')
		else:
			self.Mouth_Sht.cell(row = i,column = j).value = celval
		pass

	def cellborder(self, startrow, style = 1):
		
		if style == 0:
			thin = Side(style=None)
		else:
			thin = Side(style="thin", color="000000")
			
		for i in range(startrow, startrow + 10):
			for j in range((self.colnum_mouth + 2), (self.colnum_mouth + 5)):
				self.Mouth_Sht.cell(row = i,column = j).border = Border(top=thin, left=thin, right=thin, bottom=thin)
		pass

	def mergecells(self,startrow):										# 合并某些单元格
		self.Mouth_Sht.merge_cells(start_row=startrow, start_column=(self.colnum_mouth + 2), end_row=(startrow+9), end_column=(self.colnum_mouth + 2))
		pass

def main():
	FileOpt = cFileOpt()
	FileOpt.Prepare()
	FileOpt.ParseRawDat()
	FileOpt.CloseFiles()
	
if __name__ == '__main__':
	main()
