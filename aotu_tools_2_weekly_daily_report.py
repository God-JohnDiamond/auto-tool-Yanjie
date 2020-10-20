'''
Author: John Diamond
Date: 2020-10-19 10:12:38
LastEditors: John Diamond
LastEditTime: 2020-10-20 18:05:05
FilePath: \auto-tools\aotu_tools_2_weekly_daily_report.py
'''
# -*- coding:utf-8 -*-
import openpyxl
from datetime import date

class cFileOpt:
	NowDate = date.today()
	print('今天是：%s' % NowDate)
	def __init__(self):
		self.InputFil = openpyxl.load_workbook('2.xlsx')
		print('文件 2.xlsx 打开成功！')

	def CloseFiles(self):
		self.InputFil.save(filename = '2.xlsx')
		print('文件 2.xlsx 保存成功！' )
		self.InputFil.close()
		pass

	def Prepare(self):
		self.RawDatSht = self.InputFil.worksheets[0]		# 0 -- raw data sheet
		self.Daily_Sht = self.InputFil.worksheets[3]		# 1 -- daily report sheet		3 -- for debug
		for x in range(3, 500):
			for y in range(1, 50):
				self.Daily_Sht.cell(row = x,column = y).value = ''
				self.Daily_Sht.cell(row = x,column = y).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
				self.Daily_Sht.cell(row = x,column = y).font = openpyxl.styles.Font(name='微软雅黑', size=9)
				pass
			pass
		#self.Daily_Sht.unmerge_cells('A3:R500')
		self.WeeklySht = self.InputFil.worksheets[2]		# 2 -- weekly report sheet
		pass

	def PreFillDailySht(self, *ProjectList):
		
		pass

	def FillDailySht(self, i, j, celval):
		if celval == 0:
			celval = ''
		self.Daily_Sht.cell(row = i,column = j).value = celval
		pass

	def FillWeeklySht(self):

		pass

	def ParseRawDat(self):
		ProjectSet = set()									# 去重用的项目名称集合
		ProjectList = []									# 去重之后的项目名称
		ProjectNum = 1										# 项目序号
		RowFillDaily = 3									# 略过表头两行

		for i in range(2, self.RawDatSht.max_row+1):
			ProjectSet.add(self.RawDatSht.cell(row = i,column = 2).value)
			
		ProjectList = list(ProjectSet)						# 项目集合 ——> list
		#print(ProjectList)
		
		for j in range(len(ProjectList)):					# 项目循环
			if ProjectList[j] == None:						# 排除无内容列表
				continue
			JobSet = set()									# 去重用的岗位名称集合
			JobList = []									# 去重之后的岗位集合
			for projectloop in range(2, self.RawDatSht.max_row+1):	# 遍历以查找
				CelVal = self.RawDatSht.cell(row = projectloop,column = 2).value
				if CelVal == ProjectList[j]:				# 一个项目下
					JobSet.add(self.RawDatSht.cell(row = projectloop,column = 3).value)
			JobList = list(JobSet)
			#print(JobList)
			print('No.%d' % ProjectNum)
			print(ProjectList[j])
			self.FillDailySht(RowFillDaily, 1, ProjectNum)				# 写项目序号
			self.FillDailySht(RowFillDaily, 3, ProjectList[j])			# 写项目名
			
			for k in range(len(JobList)):								# 当前项目下的岗位数循环
				Cnt_Recommend = 0										# 推荐数
				Cnt_Effective = 0										# 有效数
				Cnt_Oneside = 0											# 一面数
				Cnt_Endface = 0											# 终面数
				Cnt_Offer = 0											# offer数
				Cnt_Entry = 0											# 入职数

				if JobList[k] == None:									# 排除无内容列表
					continue
				print('%d:%s' % (k + 1, JobList[k]))

				for jobloop in range(2, self.RawDatSht.max_row+1):		# 遍历以查找当前项目下的岗位1， 2，...
					CelVal = self.RawDatSht.cell(row = jobloop,column = 3).value
					if CelVal == JobList[k]:							# 遍历当前项目下的岗位
						Cnt_Recommend += 1								# 有符合岗位的 推荐数+1
						if self.RawDatSht.cell(row = jobloop,column = 10).value == '是':
							Cnt_Effective += 1							# 简历通过 有效数+1
							pass
						if self.RawDatSht.cell(row = jobloop,column = 12).value == '是':
							Cnt_Oneside += 1							# 简历通过 一面数+1
							pass
						if self.RawDatSht.cell(row = jobloop,column = 13).value == '是':
							Cnt_Endface += 1							# 简历通过 终面数+1
							pass
						if self.RawDatSht.cell(row = jobloop,column = 15).value == '是':
							Cnt_Offer += 1								# 简历通过 offer数+1
							updateDat = self.RawDatSht.cell(row = jobloop,column = 18).value
							print(updateDat)
							print(self.NowDate)

							pass
						if self.RawDatSht.cell(row = jobloop,column = 17).value == '是':
							Cnt_Entry += 1								# 简历通过 入职数+1
							pass
						Recomm_date = self.RawDatSht.cell(row = jobloop,column = 5).value
																		# 获取推荐日期
					pass
				# print('推荐:%d' % Cnt_Recommend)
				# print('有效:%d' % Cnt_Effective)
				# print('一面:%d' % Cnt_Oneside)
				# print('终面:%d' % Cnt_Endface)
				# print('offer:%d' % Cnt_Offer)
				# print('入职:%d' % Cnt_Entry)
				self.FillDailySht(RowFillDaily, 2, Recomm_date)			# 写推荐日期
				self.FillDailySht(RowFillDaily, 7, JobList[k])			# 写岗位名
				self.FillDailySht(RowFillDaily, 10, Cnt_Recommend)		# 写各个阶段数量
				self.FillDailySht(RowFillDaily, 11, Cnt_Effective)
				self.FillDailySht(RowFillDaily, 12, Cnt_Oneside)
				self.FillDailySht(RowFillDaily, 13, Cnt_Endface)
				self.FillDailySht(RowFillDaily, 14, Cnt_Offer)
				self.FillDailySht(RowFillDaily, 15, Cnt_Entry)
				
				RowFillDaily += 1										# 每完成一个岗位 行数+1 下移一行

																		# 合并某些单元格
				#mergecells(start_column, (RowFillDaily + len(JobList) - 1), 1)
				mergeStart = ['']
				mergeEnd = ['']
				#self.Daily_Sht.merge_cells('A3:A5')
				#self.Daily_Sht.merge_cells(start_row=RowFillDaily, start_column=1, end_row=(RowFillDaily + len(JobList) - 1), end_column=1)
				'''
				self.Daily_Sht.merge_cells(start_row=RowFillDaily, start_column=3, end_row=(RowFillDaily + len(JobList) - 1), end_column=3)
				self.Daily_Sht.merge_cells(start_row=RowFillDaily, start_column=4, end_row=(RowFillDaily + len(JobList) - 1), end_column=4)
				self.Daily_Sht.merge_cells(start_row=RowFillDaily, start_column=5, end_row=(RowFillDaily + len(JobList) - 1), end_column=5)
				self.Daily_Sht.merge_cells(start_row=RowFillDaily, start_column=6, end_row=(RowFillDaily + len(JobList) - 1), end_column=6)
				self.Daily_Sht.merge_cells(start_row=RowFillDaily, start_column=9, end_row=(RowFillDaily + len(JobList) - 1), end_column=9)
				'''
				pass
			print('\n')
			ProjectNum += 1												# 完成一个项目 序号+1
		pass
	def mergecells(self,startrow, endrow, column):
		#self.Daily_Sht.merge_cells('A3:A5')
		pass


def main():
	FileOpt = cFileOpt()
	FileOpt.Prepare()
	FileOpt.ParseRawDat()
	FileOpt.CloseFiles()
	
if __name__ == '__main__':
	main()